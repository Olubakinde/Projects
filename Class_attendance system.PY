import openpyxl  # Importing the openpyxl library for working with Excel files
from openpyxl.styles import PatternFill  # Importing PatternFill for styling cells
from datetime import datetime, timedelta  # Importing datetime and timedelta for working with dates and times
import os  # Importing the os module for interacting with the operating system
import smtplib  # Importing smtplib for sending emails
from email.mime.multipart import MIMEMultipart  # Importing MIMEMultipart for creating email messages
from email.mime.text import MIMEText  # Importing MIMEText for creating email message bodies
import matplotlib.pyplot as plt  # Importing matplotlib for data visualization

class AttendanceTaker:
    def __init__(self, class_name):
        self.class_name = class_name  # Initializing the class name
        self.student_attendance = {}  # Creating an empty dictionary to store student attendance data
        self.workbook = openpyxl.Workbook()  # Creating a new Excel workbook
        self.sheet = self.workbook.active  # Getting the active worksheet in the workbook
        self.sheet['A1'] = "Student ID"  # Setting the header for student ID column
        self.sheet['B1'] = "Student Name"  # Setting the header for student name column
        self.sheet['C1'] = "Attendance Count"  # Setting the header for attendance count column
        self.sheet['D1'] = "Swipe Time"  # Setting the header for swipe time column
        self.sheet['E1'] = "Status"  # Setting the header for status column
        self.sheet['F1'] = "Absences"  # Setting the header for absences column
        self.row = 2  # Initializing the starting row for data entry
        self.class_end_time = datetime.strptime("4:40 PM", "%I:%M %p")  # Setting the end time for the class
        self.class_start_time = datetime.strptime("2:30 PM", "%I:%M %p")  # Setting the start time for the class

    def swipe_card(self, student_id, swipe_time=None):
        if student_id not in self.student_attendance:
            self.student_attendance[student_id] = {"name": self.get_student_name(student_id), "attendance": [], "absences": 0, "late_notification_sent": False}

        if swipe_time is None:
            swipe_time = datetime.now()

        if swipe_time < self.class_start_time:
            # Send early swipe notification
            self.send_early_swipe_notification(student_id)
        elif swipe_time > self.class_end_time:
            # Send late swipe notification
            self.send_late_swipe_notification(student_id)
            self.student_attendance[student_id]["late_notification_sent"] = True
        else:
            # Record attendance
            last_swipe = self.student_attendance[student_id]["attendance"][-1]["time"] if self.student_attendance[student_id]["attendance"] else self.class_start_time
            if swipe_time - last_swipe <= timedelta(minutes=40):
                # Attendance marked present
                attendance_status = "Present"
            else:
                # Attendance marked absent
                attendance_status = "Absent"
                self.student_attendance[student_id]["absences"] += 1

            self.student_attendance[student_id]["attendance"].append({"status": attendance_status, "time": swipe_time})

            # Send email notification regardless of attendance status
            message_body = f"Dear Student,\n\nYou have swiped your card for the class on {swipe_time.strftime('%Y-%m-%d %I:%M %p')}."
            self.send_email_notification(student_id, message_body)

    def get_student_name(self, student_id):
        student_names = {
            "123456789": "John Doe",  # Mapping student IDs to their names
            "329847483": "Randy Orton",  # Additional student ID and name
        }
        return student_names.get(student_id, "Unknown")  # Returning the student's name or "Unknown" if not found

    def generate_spreadsheet(self):
        for student_id, data in self.student_attendance.items():
            student_name = data["name"]
            attendance_count = len(data["attendance"])
            absence_count = data["absences"]

            for attendance_record in data["attendance"]:
                attendance_status = attendance_record["status"]
                swipe_time = attendance_record["time"]

                self.sheet[f'A{self.row}'] = student_id
                self.sheet[f'B{self.row}'] = student_name
                self.sheet[f'C{self.row}'] = attendance_count
                self.sheet[f'D{self.row}'] = swipe_time.strftime("%Y-%m-%d %I:%M %p")
                self.sheet[f'E{self.row}'] = attendance_status
                self.sheet[f'F{self.row}'] = absence_count

                cell = self.sheet[f'E{self.row}']
                if attendance_status == "Present":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                self.row += 1

        filename = f"{self.class_name}_attendance.xlsx"
        file_path = os.path.join(os.path.expanduser('~'), 'Documents', filename)
        try:
            self.workbook.save(file_path)
            print(f"Attendance spreadsheet saved as {file_path}")
            self.visualize_attendance()
        except Exception as e:
            print(f"An error occurred: {e}")

    def send_email_notification(self, student_id, message_body):
        sender_email = "olubakindetito@gmail.com"
        receiver_email = "tolubaki@udel.edu"
        password = "cxfg alju sfta mnha"

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = f"{self.class_name} Attendance Report"

        message.attach(MIMEText(message_body, "plain"))

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            print("Email notification sent successfully.")
        except Exception as e:
            print(f"Failed to send email: {e}")

    def visualize_attendance(self):
        students = list(self.student_attendance.keys())
        absent_counts = [data["absences"] for data in self.student_attendance.values()]
        present_counts = [len(data["attendance"]) - data["absences"] for data in self.student_attendance.values()]

        labels = ['Present', 'Absent']
        sizes = [sum(present_counts), sum(absent_counts)]
        colors = ['lightgreen', 'lightcoral']

        plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=140)
        plt.axis('equal')
        plt.title('Attendance Visualization')
        plt.show()

        for student_id, count in zip(students, absent_counts):
            print(f"Student ID: {student_id}, Absent Count: {count}")

    def send_early_swipe_notification(self, student_id):
        print("Sending early swipe notification...")
        sender_email = "olubakindetito@gmail.com"
        receiver_email = "tolubaki@udel.edu"
        password = "cxfg alju sfta mnha"

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = "Early Swipe Notification"

        current_date = datetime.now().strftime("%Y-%m-%d %I:%M %p")
        body = f"Dear Student,\n\nYou swiped your card early for the class on {current_date}. Please ensure to swipe your card within the specified time frame.\n\nRegards,\nAttendance System"
        message.attach(MIMEText(body, "plain"))

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            print("Early swipe notification sent successfully.")
        except Exception as e:
            print(f"Failed to send early swipe notification: {e}")

    def send_late_swipe_notification(self, student_id):
        print("Sending late swipe notification...")
        sender_email = "olubakindetito@gmail.com"
        receiver_email = "tolubaki@udel.edu"
        password = "cxfg alju sfta mnha"

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = "Late Swipe Notification"

        current_date = datetime.now().strftime("%Y-%m-%d %I:%M %p")
        body = f"Dear Student,\n\nYou have been marked absent for the class on {current_date} due to a late swipe. Please ensure to arrive on time for the next class.\n\nRegards,\nAttendance System"
        message.attach(MIMEText(body, "plain"))

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            print("Late swipe notification sent successfully.")
        except Exception as e:
            print(f"Failed to send late swipe notification: {e}")


# Example usage:
attendance_taker = AttendanceTaker("cs181")
# User swipes card early
attendance_taker.swipe_card("123456789", datetime.strptime("2:00 PM", "%I:%M %p")) 
# User swipes card within the time frame
attendance_taker.swipe_card("123456789", datetime.strptime("2:40 PM", "%I:%M %p"))
# User swipes card after the class is over
attendance_taker.swipe_card("123456789", datetime.strptime("4:45 PM", "%I:%M %p"))
attendance_taker.generate_spreadsheet()

# Create a new instance of the AttendanceTaker class for a student who does not swipe their card
absent_student_taker = AttendanceTaker("cs181")

# Send a notification to the absent student
message_body = "Dear Student,\n\nYou were marked absent from the class today. It is important to attend classes regularly to grasp the material being taught. Please make sure to attend the next class.\n\nRegards,\nAttendance System"
absent_student_taker.send_email_notification("123456789", message_body)
