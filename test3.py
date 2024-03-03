import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import matplotlib.pyplot as plt
from collections import defaultdict

class AttendanceTaker:
    def __init__(self, class_name):
        self.class_name = class_name
        self.student_attendance = {}
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet['A1'] = "Student ID"
        self.sheet['B1'] = "Student Name"
        self.sheet['C1'] = "Attendance Count"
        self.sheet['D1'] = "Swipe Time"
        self.sheet['E1'] = "Status"
        self.sheet['F1'] = "Absences"
        self.row = 2
        self.class_end_time = datetime.strptime("16:40", "%H:%M")
        self.class_start_time = datetime.strptime("14:30", "%H:%M")

    def swipe_card(self, student_id, swipe_time=None):
        if student_id not in self.student_attendance:
            self.student_attendance[student_id] = {"name": self.get_student_name(student_id), "attendance": [], "absences": 0}

        if swipe_time is None:
            swipe_time = datetime.now()

        attendance_status = "Present" if not self.student_attendance[student_id]["attendance"] or self.student_attendance[student_id]["attendance"][-1]["status"] == "Present" else "Absent"
        self.student_attendance[student_id]["attendance"].append({"status": attendance_status, "time": swipe_time})

        if attendance_status == "Absent":
            self.student_attendance[student_id]["absences"] += 1

    def get_student_name(self, student_id):
        student_names = {
            "123456789": "John Doe"
        }
        return student_names.get(student_id, "Unknown")

    def generate_spreadsheet(self):
        for student_id, data in self.student_attendance.items():
            student_name = data["name"]
            attendance_count = len(data["attendance"])
            absence_count = data["absences"]

            for attendance_record in data["attendance"]:
                attendance_status = attendance_record["status"]
                swipe_time = attendance_record["time"]

                self.sheet[f'A{self.row}'] = student_id
                self.sheet[f'B{self.row}'] = student_name
                self.sheet[f'C{self.row}'] = attendance_count
                self.sheet[f'D{self.row}'] = swipe_time.strftime("%Y-%m-%d %H:%M:%S")
                self.sheet[f'E{self.row}'] = attendance_status
                self.sheet[f'F{self.row}'] = absence_count

                cell = self.sheet[f'E{self.row}']
                if attendance_status == "Present":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                self.row += 1

        filename = f"{self.class_name}_attendance.xlsx"
        file_path = os.path.join(os.path.expanduser('~'), 'Documents', filename)
        try:
            self.workbook.save(file_path)
            print(f"Attendance spreadsheet saved as {file_path}")
            self.send_email_notification(file_path)
            self.visualize_attendance()
        except Exception as e:
            print(f"An error occurred: {e}")

    def send_email_notification(self, file_path):
        sender_email = "olubakindetito@gmail.com"
        receiver_email = "tolubaki@udel.edu"
        password = "cxfg alju sfta mnha"

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = f"{self.class_name} Attendance Report"

        current_date = datetime.now().strftime("%Y-%m-%d")
        body = f"Subject: Attendance Report for {self.class_name} - {current_date}\n\nDear [Recipient's Name],\n\nI hope this message finds you well.\n\nThis email serves as an official attendance report for the {self.class_name} session held on {current_date}. As per our records, you were marked present for the entirety of the class.\n\nShould you have any questions or concerns regarding your attendance or require further documentation, please do not hesitate to reach out to us.\n\nThank you for your commitment to your academic pursuits. We look forward to your continued participation and engagement in future sessions.\n\nRegards,\nAttendance System"
        message.attach(MIMEText(body, "plain"))

        with open(file_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename= {os.path.basename(file_path)}",
        )
        message.attach(part)

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            print("Email notification sent successfully.")
        except Exception as e:
            print(f"Failed to send email: {e}")

    def visualize_attendance(self):
        students = list(self.student_attendance.keys())
        present_counts = [6] * len(students)
        absent_counts = [1] * len(students)

        labels = ['Present', 'Absent']
        sizes = [sum(present_counts), sum(absent_counts)]
        colors = ['lightgreen', 'lightcoral']

        plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=140)
        plt.axis('equal')
        plt.title('Attendance Visualization')
        plt.show()

        for student_id, count in zip(students, absent_counts):
            print(f"Student ID: {student_id}, Absent Count: {count}")

    def view_past_attendance(self, student_id):
        if student_id in self.student_attendance:
            print(f"Attendance History for Student: {self.student_attendance[student_id]['name']}")
            for record in self.student_attendance[student_id]['attendance']:
                print(f"Swipe Time: {record['time']}, Status: {record['status']}")
        else:
            print("No attendance records found for this student.")
    
    def update_absence(self, student_id, current_time):
        if student_id in self.student_attendance:
            last_swipe_time = self.student_attendance[student_id]["attendance"][-1]["time"]
            if current_time - last_swipe_time > timedelta(hours=2):
                self.student_attendance[student_id]["absences"] += 1
                self.send_absence_notification(student_id)
        else:
            print("No attendance records found for this student.")

    def send_absence_notification(self, student_id):
        sender_email = "olubakindetito@gmail.com"
        receiver_email = "tolubaki@udel.edu"
        password = "cxfg alju sfta mnha"

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = "Absentee Notification"

        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        body = f"Dear Student,\n\nYou have been marked absent for the class on {current_date} due to not swiping your card within the specified time frame. Please ensure to attend the next class.\n\nRegards,\nAttendance System"
        message.attach(MIMEText(body, "plain"))

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            print("Absence notification sent successfully.")
        except Exception as e:
            print(f"Failed to send absence notification: {e}")

    def check_for_late_swipe(self, student_id):
            if student_id in self.student_attendance:
                last_swipe_time = self.student_attendance[student_id]["attendance"][-1]["time"]
                if last_swipe_time > self.class_end_time and last_swipe_time < self.class_start_time:
                    self.update_absence(student_id, last_swipe_time)
                    self.send_late_swipe_notification(student_id)

    def send_late_swipe_notification(self, student_id):
        sender_email = "olubakindetito@gmail.com"
        receiver_email = "tolubaki@udel.edu"
        password = "cxfg alju sfta mnha"

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = "Late Swipe Notification"

        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        body = f"Dear Student,\n\nYou have been marked absent for the class on {current_date} due to a late swipe. Please ensure to arrive on time for the next class.\n\nRegards,\nAttendance System"
        message.attach(MIMEText(body, "plain"))

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            print("Late swipe notification sent successfully.")
        except Exception as e:
            print(f"Failed to send late swipe notification: {e}")


    def check_for_early_leave(self, student_id):
        if student_id in self.student_attendance:
            last_swipe_time = self.student_attendance[student_id]["attendance"][-1]["time"]
            if self.class_start_time - last_swipe_time < timedelta(minutes=30):
                self.send_early_leave_notification(student_id)

    def send_early_leave_notification(self, student_id):
        sender_email = "olubakindetito@gmail.com"
        receiver_email = "tolubaki@udel.edu"
        password = "cxfg alju sfta mnha"

        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = "Early Leave Notification"

        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        body = f"Dear Student,\n\nYou have been marked absent for the class on {current_date} due to leaving the class early. Please ensure to attend the full duration of the class.\n\nRegards,\nAttendance System"
        message.attach(MIMEText(body, "plain"))

        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, message.as_string())
            print("Early leave notification sent successfully.")
        except Exception as e:
            print(f"Failed to send early leave notification: {e}")

if __name__ == "__main__":
    attendance_taker = AttendanceTaker("Mathematics")

    swipe_time = datetime.now()
    for _ in range(6):
        attendance_taker.swipe_card("123456789", swipe_time=swipe_time)
    
    # Simulate a late swipe after class
    late_swipe_time = datetime.strptime("16:45", "%H:%M")
    attendance_taker.swipe_card("123456789", swipe_time=late_swipe_time)

    # Example usage to view past attendance
    attendance_taker.view_past_attendance("123456789")

    # Check for late swipes and update absences
    attendance_taker.check_for_late_swipe("123456789")

    # Check for early leave and send notifications
    attendance_taker.check_for_early_leave("123456789")

    # Generate the attendance spreadsheet after updating absences
    attendance_taker.generate_spreadsheet()
